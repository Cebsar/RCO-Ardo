import os
from datetime import datetime

from openpyxl import Workbook

from src.infrastructure.excel_adapter.workbook_reader import WorkbookReader


def create_test_workbook(path: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    # headers
    ws1.append(["Company", "AccountCode", "AccountingDate", "Debit", "Credit"])
    # data rows
    ws1.append(["ARDO", "1.1.1", datetime(2026, 1, 1), 100.0, None])
    ws1.append(["ARDO", "1.1.2", datetime(2026, 1, 2), None, 50.0])
    # merged cells
    ws1.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws1.cell(row=5, column=1).value = "Merged"
    # hidden row
    ws1.row_dimensions[4].hidden = True
    # hidden column
    ws1.column_dimensions['E'].hidden = True

    # named range
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName('TestRange', attr_text=f"'{ws1.title}'!$A$1:$B$2")
    wb.defined_names.append(dn)

    # second sheet
    ws2 = wb.create_sheet(title="Sheet2")
    ws2.append([None, None])
    wb.save(path)


def test_excel_adapter_reads_rows(tmp_path):
    f = tmp_path / "test_master.xlsx"
    create_test_workbook(str(f))

    reader = WorkbookReader(read_only=False)
    meta = reader.load(str(f))

    assert "Sheet1" in meta.sheet_names
    headers = reader.get_headers("Sheet1")
    assert headers[0] == "Company"

    rows = list(reader.get_rows("Sheet1"))
    # two data rows expected (hidden row skipped)
    assert any(r.get("Company") == "ARDO" for r in rows)
    assert len(rows) >= 2
