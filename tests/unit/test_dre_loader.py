from datetime import date
from pathlib import Path
import tempfile

from openpyxl import Workbook

from src.infrastructure.dre_loader.builder import DRETreeBuilder
from src.infrastructure.dre_loader.reader import OverviewReader
from src.infrastructure.dre_loader.parser import HierarchyParser
from src.infrastructure.dre_loader.models import HierarchyReport


def test_overview_reader_reads_rows():
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "overview.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Overview RCO"
        ws.append(["N1", "N2", "N3"])
        ws.append(["Revenue", None, None])
        ws.append([None, "Sales", None])
        ws.append([None, None, "Online"])
        wb.save(path)

        reader = OverviewReader(path=path)
        rows, report = reader.read()

        assert report.rows_read == 4
        assert rows[0]["col_1"] == "N1"
        assert rows[1]["col_1"] == "Revenue"


def test_overview_reader_preserves_hidden_rows_and_formula_metadata():
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "overview.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Overview RCO"
        ws.append(["N1", "Real"])
        ws.append(["Revenue", "=SUM(B3:B4)"])
        ws.row_dimensions[2].hidden = True
        wb.save(path)

        reader = OverviewReader(path=path)
        rows, _ = reader.read()

        assert rows[1]["_hidden"] is True
        assert rows[1]["_formulas"] == {"col_2": "=SUM(B3:B4)"}


def test_hierarchy_parser_detects_levels():
    rows = [
        {"col_1": "N1", "col_2": "N2", "col_3": "N3", "_row_number": 1},
        {"col_1": "Revenue", "col_2": None, "col_3": None, "_row_number": 2},
        {"col_1": None, "col_2": "Sales", "col_3": None, "_row_number": 3},
        {"col_1": None, "col_2": None, "col_3": "Online", "_row_number": 4},
    ]
    parser = HierarchyParser(rows)
    items = parser.parse()

    assert len(items) == 3
    assert items[0].label == "Revenue" and items[0].level == 1
    assert items[1].label == "Sales" and items[1].level == 2
    assert items[2].label == "Online" and items[2].level == 3


def test_hierarchy_parser_classifies_empty_separator_hidden_and_account_rows():
    rows = [
        {"col_1": "N1", "col_2": "N2", "_row_number": 1},
        {"col_1": None, "col_2": None, "_row_number": 2, "_hidden": True},
        {"col_1": "-----", "col_2": None, "_row_number": 3},
        {"col_1": "1.1.2.01.9999 - Clientes", "col_2": None, "_row_number": 4},
        {"col_1": "Total Receita", "col_2": None, "_row_number": 5, "_formulas": {"col_3": "=SUM(C1:C4)"}},
    ]
    parser = HierarchyParser(rows)
    items = parser.parse()

    assert parser.row_classifications[2] == "structural empty row"
    assert parser.row_classifications[3] == "visual separator"
    assert parser.row_classifications[4] == "analytical account"
    assert parser.row_classifications[5] == "synthetic account"
    assert [item.label for item in items] == ["1.1.2.01.9999 - Clientes", "Total Receita"]


def test_dre_tree_builder_preserves_hierarchy():
    rows = [
        {"col_1": "N1", "col_2": "N2", "col_3": "N3", "_row_number": 1},
        {"col_1": "Revenue", "col_2": None, "col_3": None, "_row_number": 2},
        {"col_1": None, "col_2": "Sales", "col_3": None, "_row_number": 3},
        {"col_1": None, "col_2": None, "col_3": "Online", "_row_number": 4},
    ]

    builder = DRETreeBuilder()
    tree, report = builder.build_from_rows(rows)

    assert report.nodes_count == 3
    assert report.root_nodes == 1
    assert len(tree.roots) == 1
    root = tree.roots[0]
    assert root.name == "Revenue"
    assert len(root.children) == 1
    assert root.children[0].name == "Sales"
    assert len(root.children[0].children) == 1
    assert root.children[0].children[0].name == "Online"
    assert report.tree_validation is not None
    assert report.tree_validation.is_valid
    assert report.tree_validation.parent_by_code == {
        "Revenue": None,
        "Sales": "Revenue",
        "Online": "Sales",
    }
    assert report.traversal is not None
    assert report.traversal.traversal_order == ("Revenue", "Sales", "Online")
    assert report.execution_metrics is not None
    assert report.execution_metrics.nodes_processed == 3
    assert report.execution_metrics.max_depth == 3


def test_dre_tree_builder_promotes_orphans_deterministically():
    rows = [
        {"col_1": "N1", "col_2": "N2", "col_3": "N3", "_row_number": 1},
        {"col_1": None, "col_2": "Sales", "col_3": None, "_row_number": 2},
        {"col_1": "Revenue", "col_2": None, "col_3": None, "_row_number": 3},
    ]

    builder = DRETreeBuilder()
    tree, report = builder.build_from_rows(rows)

    assert [node.name for node in tree.roots] == ["Sales", "Revenue"]
    assert report.warnings == ["Row 2 has level 2 without parent; promoting to root"]
    assert report.tree_validation is not None
    assert not report.tree_validation.is_valid
    assert report.traversal is not None
    assert report.traversal.traversal_order == ("Sales", "Revenue")


def test_dre_tree_builder_from_path_reads_source():
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "overview.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Overview RCO"
        ws.append(["N1", "N2", "N3"])
        ws.append(["Revenue", None, None])
        ws.append([None, "Sales", None])
        wb.save(path)

        builder = DRETreeBuilder()
        tree, report = builder.build_from_path(path)

        assert report.rows_read == 3
        assert len(tree.roots) == 1
        assert tree.metadata["sheet_name"] == "Overview RCO"


def test_dre_tree_builder_reports_row_classification_metadata():
    rows = [
        {"col_1": "N1", "_row_number": 1},
        {"col_1": None, "_row_number": 2, "_hidden": True},
        {"col_1": "Revenue", "_row_number": 3, "_formulas": {"col_2": "=SUM(B:B)"}},
    ]

    builder = DRETreeBuilder()
    tree, report = builder.build_from_rows(rows)

    assert report.metadata["row_classifications"][2] == "structural empty row"
    assert report.metadata["hidden_rows"] == [2]
    assert report.metadata["formula_rows"] == [3]
    assert tree.metadata["row_classifications"][3] == "synthetic account"
