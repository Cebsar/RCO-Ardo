import os
import tempfile
from datetime import datetime

from openpyxl import Workbook

from src.metadata_engine.extractor import MetadataExtractor


def create_test_workbook(path: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    # headers
    ws1.append(["Company", "AccountCode", "AccountingDate", "Debit", "Credit"])
    # data rows
    ws1.append(["ARDO", "1.1.1", datetime(2026, 1, 1), 100.0, None])
    ws1.append(["ARDO", "1.1.2", datetime(2026, 1, 2), None, 50.0])
    # merged cells
    ws1.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws1.cell(row=5, column=1).value = "Merged"
    # hidden row
    ws1.row_dimensions[4].hidden = True
    # hidden column
    ws1.column_dimensions['E'].hidden = True

    # named range (defined name)
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName('TestRange', attr_text=f"'{ws1.title}'!$A$1:$B$2")
    wb.defined_names.add(dn)

    # second sheet
    ws2 = wb.create_sheet(title="Sheet2")
    ws2.append([None, None])
    wb.save(path)


def test_metadata_extraction(tmp_path):
    f = tmp_path / "test_master.xlsx"
    create_test_workbook(str(f))

    extractor = MetadataExtractor(read_only=False)
    meta = extractor.extract(str(f))

    assert meta.path == str(f)
    assert "Sheet1" in meta.sheet_names
    assert "Sheet2" in meta.sheet_names
    # worksheets metadata
    sheet1 = next(w for w in meta.worksheets if w.name == "Sheet1")
    assert sheet1.max_row >= 5
    assert any(h == "Company" for h in sheet1.headers if h)
    assert len(sheet1.merged_cells) >= 1
    assert 4 in sheet1.hidden_rows or 5 in sheet1.hidden_rows
    # named ranges
    assert any(n.name == "TestRange" for n in meta.named_ranges)
    assert meta.execution_time_seconds >= 0.0
