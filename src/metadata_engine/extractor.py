from __future__ import annotations

import logging
from time import perf_counter
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .models import (
    WorkbookMetadata,
    WorksheetMetadata,
    ColumnMetadata,
    NamedRangeMetadata,
)

logger = logging.getLogger(__name__)


class MetadataExtractor:
    """Extract metadata from an Excel workbook.

    Responsibilities (metadata-only):
    - open workbook
    - detect version (openpyxl version)
    - list sheets
    - detect used ranges
    - detect headers (first non-empty row)
    - detect merged cells
    - detect hidden rows and columns
    - detect named ranges
    - return structured metadata objects
    """

    def __init__(self, read_only: bool = True):
        self.read_only = read_only

    def extract(self, path: str) -> WorkbookMetadata:
        logger.info("Starting metadata extraction for %s", path)
        start = perf_counter()
        wb = self._open_workbook(path)
        try:
            meta = self._extract_from_workbook(path, wb)
            return meta
        finally:
            execution_time = perf_counter() - start
            logger.info("Metadata extraction completed in %.4f seconds", execution_time)
            meta.execution_time_seconds = execution_time

    def _open_workbook(self, path: str) -> Workbook:
        logger.debug("Loading workbook: %s (read_only=%s)", path, self.read_only)
        wb = load_workbook(filename=path, read_only=self.read_only, data_only=True)
        return wb

    def _extract_from_workbook(self, path: str, wb: Workbook) -> WorkbookMetadata:
        engine_version = getattr(wb, "excel_base_date", None)
        # excel_base_date is not a version; better to include openpyxl version
        try:
            import openpyxl

            openpyxl_version = getattr(openpyxl, "__version__", None)
        except Exception:
            openpyxl_version = None

        sheet_names = list(wb.sheetnames)
        named_ranges = self._extract_named_ranges(wb)
        worksheets = [self._extract_worksheet(wb[s]) for s in sheet_names]

        return WorkbookMetadata(
            path=path,
            engine_version=openpyxl_version,
            sheet_names=sheet_names,
            worksheets=worksheets,
            named_ranges=named_ranges,
            execution_time_seconds=0.0,
        )

    def _extract_named_ranges(self, wb: Workbook) -> List[NamedRangeMetadata]:
        nrs: List[NamedRangeMetadata] = []
        defined_names = wb.defined_names
        names = defined_names.values() if hasattr(defined_names, "values") else defined_names
        for name in names:  # type: ignore
            # name is an openpyxl.workbook.defined_name.DefinedName
            attr_text = getattr(name, "attr_text", None)
            if attr_text is None:
                continue
            # destinations yields (sheetName, range) tuples
            sheet = None
            rng = attr_text
            try:
                destinations = list(name.destinations)
                if destinations:
                    sheet, rng = destinations[0]
            except Exception:
                # some defined names may not expose destinations cleanly
                pass
            nrs.append(NamedRangeMetadata(name=name.name, sheet=sheet, range=rng))
        return nrs

    def _extract_worksheet(self, ws) -> WorksheetMetadata:
        title = ws.title
        logger.debug("Extracting worksheet: %s", title)
        max_row = getattr(ws, "max_row", 0) or 0
        max_column = getattr(ws, "max_column", 0) or 0
        used_range = self._detect_used_range(ws)
        headers = self._detect_headers(ws)
        merged = [str(rng) for rng in ws.merged_cells.ranges]
        hidden_rows = [r for r, v in ws.row_dimensions.items() if getattr(v, "hidden", False)]
        # column_dimensions keys are letter-based (e.g. 'A') — convert to indices
        from openpyxl.utils import column_index_from_string

        hidden_columns_letters = [c for c, v in ws.column_dimensions.items() if getattr(v, "hidden", False)]
        hidden_columns = []
        for col in hidden_columns_letters:
            try:
                hidden_columns.append(column_index_from_string(col))
            except Exception:
                # fallback: ignore non-standard keys
                continue
        columns = []
        for idx in range(1, max_column + 1):
            header = headers[idx - 1] if idx - 1 < len(headers) else None
            is_hidden = idx in hidden_columns
            merged_range = self._find_merged_for_column(idx, ws)
            columns.append(ColumnMetadata(header=header, index=idx, is_hidden=is_hidden, merged_range=merged_range))

        return WorksheetMetadata(
            name=title,
            title=title,
            max_row=max_row,
            max_column=max_column,
            used_range=used_range,
            headers=headers,
            columns=columns,
            merged_cells=merged,
            hidden_rows=hidden_rows,
            hidden_columns=hidden_columns,
        )

    def _detect_used_range(self, ws) -> Optional[Tuple[int, int, int, int]]:
        # openpyxl does not provide used range directly; infer from data
        min_row = None
        min_col = None
        max_row = 0
        max_col = 0
        for row in ws.iter_rows(values_only=True):
            max_row += 1
            for col_idx, cell in enumerate(row, start=1):
                if cell is not None:
                    if min_row is None:
                        min_row = max_row
                    if min_col is None or col_idx < min_col:
                        min_col = col_idx
                    if col_idx > max_col:
                        max_col = col_idx
        if min_row is None:
            return None
        return (min_row, min_col or 1, max_row, max_col)

    def _detect_headers(self, ws) -> List[Optional[str]]:
        # choose first non-empty row as headers
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                return [str(cell).strip() if cell is not None else None for cell in row]
        return []

    def _find_merged_for_column(self, col_idx: int, ws) -> Optional[str]:
        for rng in ws.merged_cells.ranges:
            min_col = rng.min_col
            max_col = rng.max_col
            if min_col <= col_idx <= max_col:
                return str(rng)
        return None
