from __future__ import annotations

import logging
from time import perf_counter
from typing import Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.metadata_engine.extractor import MetadataExtractor
from src.metadata_engine.models import WorkbookMetadata
from .worksheet_reader import WorksheetReader

logger = logging.getLogger(__name__)


class WorkbookReader:
    """Reads workbook and exposes metadata and worksheet readers.

    Integrates with MetadataExtractor to produce `WorkbookMetadata` and uses
    WorksheetReader to iterate rows as raw dicts.
    """

    def __init__(self, read_only: bool = True, logger: Optional[logging.Logger] = None):
        self.read_only = read_only
        self._wb: Optional[Workbook] = None
        self._path: Optional[str] = None
        self._metadata: Optional[WorkbookMetadata] = None
        self._logger = logger or logging.getLogger(__name__)

    def load(self, path: str) -> WorkbookMetadata:
        self._path = path
        self._logger.info("Loading workbook and extracting metadata: %s", path)
        start = perf_counter()
        # extract metadata (uses openpyxl internally)
        extractor = MetadataExtractor(read_only=self.read_only)
        meta = extractor.extract(path)
        # also keep an open workbook handle for row access (read-only)
        self._wb = load_workbook(filename=path, read_only=True, data_only=True)
        self._metadata = meta
        duration = perf_counter() - start
        self._logger.info("Workbook loaded in %.4f seconds", duration)
        return meta

    def get_metadata(self) -> WorkbookMetadata:
        if self._metadata is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        return self._metadata

    def get_sheet_names(self) -> List[str]:
        return list(self.get_metadata().sheet_names)

    def get_worksheet(self, name: Optional[str] = None) -> WorksheetReader:
        if self._wb is None:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        meta = self.get_metadata()
        sheet_name = name or (meta.sheet_names[0] if meta.sheet_names else None)
        if sheet_name is None:
            raise RuntimeError("No sheet available in workbook")
        ws = self._wb[sheet_name]
        # find worksheet metadata
        ws_meta = next((w for w in meta.worksheets if w.name == sheet_name), None)
        return WorksheetReader(ws, ws_meta, logger=self._logger)

    def get_headers(self, sheet_name: Optional[str] = None) -> List[Optional[str]]:
        return self.get_worksheet(sheet_name).get_headers()

    def get_rows(self, sheet_name: Optional[str] = None) -> Iterable[dict]:
        reader = self.get_worksheet(sheet_name)
        start = perf_counter()
        for row in reader.iter_rows():
            yield row
        duration = perf_counter() - start
        self._logger.info("Iterated rows in %.4f seconds", duration)
