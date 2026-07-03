from __future__ import annotations

import logging
from typing import Any, Dict, Generator, List, Optional
from time import perf_counter

from openpyxl.worksheet.worksheet import Worksheet

from src.metadata_engine.models import WorksheetMetadata
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class RowIterator:
    def __init__(self, worksheet: Worksheet, metadata: Optional[WorksheetMetadata] = None, logger: Optional[logging.Logger] = None):
        self._ws = worksheet
        self._metadata = metadata
        self._logger = logger or logger
        self._headers: List[Optional[str]] = self._metadata.headers if (self._metadata and self._metadata.headers) else []

    def __iter__(self) -> Generator[Dict[str, Any], None, None]:
        # determine bounds
        min_row = 1
        max_row = getattr(self._ws, "max_row", 0) or 0
        min_col = 1
        max_col = getattr(self._ws, "max_column", 0) or 0

        # prefer used_range if available
        if self._metadata and self._metadata.used_range:
            min_row, min_col, used_max_row, used_max_col = self._metadata.used_range
            if used_max_row:
                max_row = used_max_row
            if used_max_col:
                max_col = used_max_col

        # ensure headers length
        if not self._headers:
            # attempt to read headers from sheet via first non-empty row
            for row in self._ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    self._headers = [str(cell).strip() if cell is not None else None for cell in row]
                    break

        # produce rows starting after header row
        header_row = 1
        # if metadata used_range indicates min_row > 1, adjust
        if self._metadata and self._metadata.used_range:
            header_row = self._metadata.used_range[0]

        start_row = header_row + 1

        start_time = perf_counter()
        for r_idx, row in enumerate(self._ws.iter_rows(min_row=start_row, max_row=max_row, values_only=False), start=start_row):
            # skip hidden rows
            if self._metadata and r_idx in (self._metadata.hidden_rows or []):
                continue
            row_dict: Dict[str, Any] = {}
            for c_idx, cell in enumerate(row, start=1):
                # skip hidden columns
                if self._metadata and c_idx in (self._metadata.hidden_columns or []):
                    continue
                header = None
                if c_idx - 1 < len(self._headers):
                    header = self._headers[c_idx - 1]
                else:
                    # fallback to column letter
                    header = get_column_letter(c_idx)
                # preserve original value
                value = None if cell is None else getattr(cell, "value", None)
                row_dict[header or get_column_letter(c_idx)] = value
            yield row_dict

        duration = perf_counter() - start_time
        self._logger.info("Row iteration completed in %.4f seconds for sheet %s", duration, getattr(self._ws, "title", "<unknown>"))

    # make class iterable
    def __next__(self):
        raise StopIteration
