from __future__ import annotations

from typing import Iterable, List
from openpyxl.worksheet.worksheet import Worksheet


class WorksheetReader:
    """Thin wrapper around openpyxl Worksheet to expose header and row iteration.

    Does not perform transformations; returns raw cell objects.
    """

    def __init__(self, worksheet: Worksheet):
        self._ws = worksheet

    def read_headers(self) -> List[str]:
        # Return first row values (may include None)
        for row in self._ws.iter_rows(min_row=1, max_row=1, values_only=False):
            return [cell.value for cell in row]
        return []

    def iter_rows(self, min_row: int = 2) -> Iterable:
        for row in self._ws.iter_rows(min_row=min_row, values_only=False):
            yield row
