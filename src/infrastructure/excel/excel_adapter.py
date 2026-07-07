from __future__ import annotations

from typing import Any, Dict, Generator, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.application.contracts.datasource import IDataSource
from .worksheet_reader import WorksheetReader
from .header_mapper import map_headers
from .cell_reader import read_cell
from .file_validator import validate as validate_headers


class ExcelAdapter(IDataSource):
    """Infrastructure adapter that reads Excel workbooks and yields raw rows as dicts.

    Public API:
    - load(path)
    - get_sheet(name=None) -> sheet name
    - get_headers(sheet_name=None) -> List[str]
    - get_rows(sheet_name=None) -> Generator[Dict[str, Any], None, None]
    """

    def __init__(self, path: Optional[str] = None):
        self.path = path
        self._wb = None

    # IDataSource implementation
    def read(self, source: str, sheet_name: Optional[str] = None) -> Any:
        self.load(source)
        return list(self.get_rows(sheet_name=sheet_name))

    def load(self, path: Optional[str] = None) -> None:
        path = path or self.path
        if not path:
            raise ValueError("Path must be provided to load workbook")
        self._wb = load_workbook(filename=path, read_only=True, data_only=True)

    def get_sheet(self, name: Optional[str] = None) -> str:
        if not self._wb:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        if name:
            if name not in self._wb.sheetnames:
                raise KeyError(f"Sheet '{name}' not found in workbook")
            return name
        return self._wb.active.title

    def get_headers(self, sheet_name: Optional[str] = None) -> List[str]:
        ws = self._get_worksheet(sheet_name)
        reader = WorksheetReader(ws)
        headers = reader.read_headers()
        return map_headers(headers)

    def get_rows(self, sheet_name: Optional[str] = None) -> Generator[Dict[str, Any], None, None]:
        ws = self._get_worksheet(sheet_name)
        reader = WorksheetReader(ws)
        headers = map_headers(reader.read_headers())
        # validate headers (returns a DTO-like object from file_validator)
        validation = validate_headers(headers)
        if validation.missing_columns or validation.duplicated_columns:
            # Adapter does not correct data; raise to let caller handle
            raise ValueError(
                f"Invalid headers: missing={validation.missing_columns} duplicated={validation.duplicated_columns}"
            )
        for row_cells in reader.iter_rows(min_row=2):
            # yield raw dictionary mapping header->cell.value
            row_dict: Dict[str, Any] = {"__row__": row_cells[0].row if row_cells else None}
            for header, cell in zip(headers, row_cells):
                row_dict[header] = read_cell(cell)
            yield row_dict

    def _get_worksheet(self, sheet_name: Optional[str]) -> Worksheet:
        if not self._wb:
            raise RuntimeError("Workbook not loaded. Call load() first.")
        name = sheet_name or self._wb.active.title
        return self._wb[name]
