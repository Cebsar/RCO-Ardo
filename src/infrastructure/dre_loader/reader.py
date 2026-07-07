from __future__ import annotations

import logging
from pathlib import Path
from time import perf_counter
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from .models import HierarchyReport

logger = logging.getLogger(__name__)


class OverviewReader:
    def __init__(self, path: Path, sheet_name: str = "Overview RCO"):
        self.path = Path(path)
        self.sheet_name = sheet_name

    def read(self) -> Tuple[List[Dict[str, Any]], HierarchyReport]:
        start = perf_counter()
        workbook = load_workbook(filename=self.path, data_only=True, read_only=False)
        formula_workbook = load_workbook(filename=self.path, data_only=False, read_only=False)
        try:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{self.sheet_name}' not found in workbook")

            worksheet = workbook[self.sheet_name]
            formula_worksheet = formula_workbook[self.sheet_name]
            rows: List[Dict[str, Any]] = []

            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                row_data = {f"col_{col_index}": cell for col_index, cell in enumerate(row, start=1)}
                row_data["_row_number"] = row_index
                row_data["_hidden"] = bool(worksheet.row_dimensions[row_index].hidden)
                formulas: Dict[str, str] = {}
                for col_index in range(1, worksheet.max_column + 1):
                    formula_value = formula_worksheet.cell(row=row_index, column=col_index).value
                    if isinstance(formula_value, str) and formula_value.startswith("="):
                        formulas[f"col_{col_index}"] = formula_value
                if formulas:
                    row_data["_formulas"] = formulas
                rows.append(row_data)
        finally:
            workbook.close()
            formula_workbook.close()

        report = HierarchyReport(rows_read=len(rows), execution_time_seconds=perf_counter() - start)
        logger.info("OverviewReader read %d rows from %s in %.6f seconds", report.rows_read, self.sheet_name, report.execution_time_seconds)
        return rows, report
