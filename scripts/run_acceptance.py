from __future__ import annotations

import hashlib
import json
import sys
from dataclasses import asdict, is_dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from time import perf_counter
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.execution.configuration import ExecutionConfiguration
from src.execution.runner import PipelineRunner
from src.infrastructure.business_rules.provider import BusinessRuleProvider
from src.infrastructure.dre_loader.builder import DRETreeBuilder
from src.infrastructure.dre_loader.parser import HierarchyParser
from src.infrastructure.dre_loader.reader import OverviewReader
from src.infrastructure.excel.excel_adapter import ExcelAdapter
from src.infrastructure.header_mapper.mapper import HeaderMapper
from src.infrastructure.reconciliation.engine import ReconciliationEngine
from src.infrastructure.rule_engine.engine import RuleEngine
from src.infrastructure.schema_validator.validator import SchemaValidator
from src.infrastructure.warehouse.builder import WarehouseBuilder


SOURCE = ROOT / "data" / "master" / "Demonstrativos_Financeiros_2026_Rel_Razao_Active.xlsx"
RULES = ROOT / "config" / "business_rules.yaml"
OUTPUT_DIR = ROOT / "outputs" / "acceptance" / "official_accounting"
OUTPUT_FILES = [
    "FACT_ACCOUNTING_ENTRY.xlsx",
    "FATO_DRE.xlsx",
    "ODS.xlsx",
    "Warehouse.xlsx",
    "Calculated_DRE.xlsx",
    "Reconciliation_Report.xlsx",
    "Validation_Report.xlsx",
    "Executive_Summary.pdf",
    "Execution_Metrics.json",
    "Validation_Report.md",
]


def decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip().replace("\u00a0", "").replace(" ", "")
        if not text:
            return None
        if text.count(",") and not text.count("."):
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def serialise(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if is_dataclass(value):
        return serialise(asdict(value))
    if isinstance(value, dict):
        return {str(k): serialise(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [serialise(v) for v in value]
    return value


def workbook_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_entry_id(index: int) -> str:
    return f"AE-{index:08d}"


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column[:200]:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 48)


def save_workbook(path: Path, sheets: dict[str, list[list[Any]]]) -> None:
    wb = Workbook()
    first = True
    for title, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        for row in rows:
            ws.append([serialise(v) for v in row])
        if rows:
            style_sheet(ws)
    wb.save(path)
    wb.close()


def run_pipeline() -> tuple[Any, WarehouseBuilder, BusinessRuleProvider, float]:
    started = perf_counter()
    warehouse_builder = WarehouseBuilder()
    business_rule_provider = BusinessRuleProvider(RULES)
    runner = PipelineRunner(
        excel_adapter=ExcelAdapter(path=str(SOURCE)),
        header_mapper=HeaderMapper(),
        schema_validator=SchemaValidator(),
        warehouse_builder=warehouse_builder,
        dre_tree_builder=DRETreeBuilder(),
        business_rule_provider=business_rule_provider,
        rule_engine=RuleEngine(),
        reconciliation_engine=ReconciliationEngine(),
    )
    report = runner.run(
        ExecutionConfiguration(
            source_path=SOURCE,
            rules_config_path=RULES,
            dre_sheet_name="Overview RCO",
        )
    )
    return report, warehouse_builder, business_rule_provider, perf_counter() - started


def dimension_rows(rows: Iterable[Any]) -> list[list[Any]]:
    output = [["surrogate_key", "natural_key", "attributes"]]
    for row in sorted(rows, key=lambda item: item.surrogate_key):
        output.append([row.surrogate_key, row.natural_key, json.dumps(serialise(row.attributes), ensure_ascii=False)])
    return output


def structural_rows(report: Any) -> list[list[Any]]:
    normalization = report.metadata["workflow_payload"].get("normalization_report")
    output = [["row_number", "classification", "reason"]]
    for issue in getattr(normalization, "issues", []):
        if issue.code == "EMPTY_ROW":
            output.append([issue.row_number, "empty structural row", issue.message])
    return output


def fact_rows(facts: Iterable[Any]) -> list[list[Any]]:
    output = [[
        "stable_entry_id",
        "surrogate_key",
        "company_key",
        "division_key",
        "costcenter_key",
        "account_key",
        "period_key",
        "account_code",
        "amount",
        "entry_type",
        "accounting_date",
        "description",
        "company",
        "division",
        "cost_center",
    ]]
    for index, fact in enumerate(sorted(facts, key=lambda item: item.surrogate_key), start=1):
        output.append([
            stable_entry_id(index),
            fact.surrogate_key,
            fact.company_key,
            fact.division_key,
            fact.costcenter_key,
            fact.account_key,
            fact.period_key,
            fact.source_row.get("account_code"),
            fact.amount,
            fact.entry_type,
            fact.accounting_date,
            fact.description,
            fact.source_row.get("company"),
            fact.source_row.get("division"),
            fact.source_row.get("cost_center"),
        ])
    return output


def flatten_dre_nodes(roots: Iterable[Any]) -> list[Any]:
    nodes: list[Any] = []

    def walk(node: Any, path: str) -> None:
        nodes.append((node, path))
        for child in node.children:
            walk(child, f"{path} > {child.name}")

    for root in roots:
        walk(root, root.name)
    return nodes


def overview_expected_values() -> dict[str, dict[str, Any]]:
    reader = OverviewReader(SOURCE, "Overview RCO")
    rows, _ = reader.read()
    parser = HierarchyParser(rows)
    items = parser.parse()

    real_column = 9
    header_row_number = 0
    for row in rows:
        for key, value in row.items():
            if value == "Real" and key.startswith("col_"):
                real_column = int(key.split("_")[1])
                header_row_number = int(row["_row_number"])
                break
        if header_row_number:
            break

    by_row = {int(row["_row_number"]): row for row in rows}
    expected: dict[str, dict[str, Any]] = {}
    for item in items:
        if item.row_number <= header_row_number:
            continue
        row = by_row.get(item.row_number, {})
        value = decimal_or_none(row.get(f"col_{real_column}"))
        expected[item.code] = {
            "row_number": item.row_number,
            "node_code": item.code,
            "node_name": item.label,
            "level": item.level,
            "expected_value": value,
        }
    return expected


def calculated_dre_rows(report: Any, business_rule_provider: BusinessRuleProvider) -> tuple[list[list[Any]], dict[str, Decimal | None]]:
    rule_execution = report.rule_execution
    result_by_code = {result.node_code: result for result in getattr(rule_execution.report, "results", [])}
    payload = report.metadata.get("workflow_payload", {})
    dre_tree = payload.get("dre_tree")
    rows = [["node_code", "node_name", "level", "path", "calculated_value", "rule_id", "matched_fact_count", "warnings", "errors"]]
    actual_by_code: dict[str, Decimal | None] = {}
    for node, path in flatten_dre_nodes(dre_tree.roots):
        result = result_by_code.get(node.code.value)
        value = result.value if result else None
        rule = business_rule_provider.get_rule_by_node(node.code.value)
        actual_by_code[node.code.value] = value
        rows.append([
            node.code.value,
            node.name,
            int(node.level),
            path,
            value,
            rule.id if rule else None,
            result.matched_fact_count if result else None,
            "; ".join(result.warnings) if result else None,
            "; ".join(result.errors) if result else None,
        ])
    return rows, actual_by_code


def difference_rows(expected: dict[str, dict[str, Any]], actual: dict[str, Decimal | None], provider: BusinessRuleProvider) -> list[list[Any]]:
    rows = [["company", "division", "period", "node", "expected_value", "calculated_value", "difference", "status", "probable_business_rule"]]
    for code, item in expected.items():
        expected_value = item["expected_value"]
        if expected_value is None:
            continue
        calculated = actual.get(code)
        difference = calculated - expected_value if calculated is not None else None
        rule = provider.get_rule_by_node(code)
        status = "PASS" if difference == Decimal("0") else "FAIL"
        rows.append([
            "ALL",
            "ALL",
            "OVERVIEW_CURRENT",
            f"{code} - {item['node_name']}",
            expected_value,
            calculated,
            difference,
            status,
            rule.id if rule else "Official Overview value",
        ])
    return rows


def count_dre_differences(dre_differences: list[list[Any]]) -> int:
    if len(dre_differences) <= 1:
        return 0
    header = dre_differences[0]
    status_index = header.index("status")
    return sum(1 for row in dre_differences[1:] if row[status_index] != "PASS")


def reconciliation_rows(report: Any, dre_differences: list[list[Any]]) -> list[list[Any]]:
    rows = [["metric", "value"]]
    audit = report.reconciliation
    rows.extend([
        ["source", audit.source],
        ["engine_nodes_compared", audit.validation.nodes_compared],
        ["engine_mismatches", audit.reconciliation.mismatches],
        ["overview_value_differences", count_dre_differences(dre_differences)],
        ["engine_duration_seconds", audit.execution.duration_seconds],
    ])
    rows.append([])
    rows.append(["company", "division", "period", "node", "expected_value", "calculated_value", "difference", "status", "probable_business_rule"])
    rows.extend(dre_differences[1:])
    return rows


def fato_dre_rows(report: Any, provider: BusinessRuleProvider) -> list[list[Any]]:
    payload = report.metadata.get("workflow_payload", {})
    dre_tree = payload.get("dre_tree")
    rule_execution = report.rule_execution
    result_by_code = {result.node_code: result for result in getattr(rule_execution.report, "results", [])}
    overview_filters = overview_filter_values()
    rows = [[
        "company",
        "division",
        "cost_center",
        "period",
        "dre_node",
        "planned_value",
        "actual_value",
        "difference",
        "percentage",
    ]]

    if dre_tree is None:
        return rows

    for node, _path in flatten_dre_nodes(dre_tree.roots):
        result = result_by_code.get(node.code.value)
        actual_value = result.value if result and result.value is not None else (node.amount.amount if node.amount is not None else None)
        rows.append([
            overview_filters.get("company"),
            overview_filters.get("division"),
            overview_filters.get("cost_center"),
            overview_filters.get("period"),
            f"{node.code.value} - {node.name}",
            None,
            actual_value,
            Decimal("0") if actual_value is not None else None,
            None,
        ])
    return rows


def _matches_filter(fact: dict[str, Any], key: str, value: Any) -> bool:
    if key.endswith("_prefix"):
        field = key[: -len("_prefix")]
        return isinstance(fact.get(field), str) and fact[field].startswith(value)
    return fact.get(key) == value


def overview_filter_values() -> dict[str, str]:
    reader = OverviewReader(SOURCE, "Overview RCO")
    rows, _ = reader.read()
    filters: dict[str, str] = {
        "company": "ALL",
        "division": "ALL",
        "cost_center": "ALL",
        "period": "OVERVIEW_CURRENT",
    }
    labels = {
        "Empresa:": "company",
        "Divisão:": "division",
        "Centro de Custo:": "cost_center",
        "Acumulado Até:": "period",
    }
    for row in rows[:8]:
        for key, label in labels.items():
            values = list(row.values())
            if key in values:
                index = values.index(key)
                if index + 2 < len(values) and values[index + 2] not in (None, ""):
                    filters[label] = str(values[index + 2])
    return filters


def stage_rows(report: Any) -> list[list[Any]]:
    rows = [["stage", "success", "duration_seconds", "errors", "warnings"]]
    for stage in report.stage_results:
        rows.append([
            stage.name,
            stage.success,
            stage.duration_seconds,
            "; ".join(stage.errors),
            "; ".join(stage.warnings),
        ])
    return rows


def validation_markdown(report: Any, differences: list[list[Any]], source_hash: str) -> str:
    normalization = report.metadata["workflow_payload"].get("normalization_report")
    warehouse = report.metadata["workflow_payload"].get("warehouse_report")
    lines = [
        "# Acceptance Validation Report",
        "",
        f"- Official workbook: `{SOURCE.relative_to(ROOT)}`",
        f"- Workbook SHA256: `{source_hash}`",
        f"- Pipeline success: `{report.success}`",
        f"- Stages passed: `{report.passed_stages}`",
        f"- Stages failed: `{report.failed_stages}`",
        f"- Normalized entries: `{getattr(normalization, 'normalized_count', 0)}`",
        f"- Invalid rows: `{getattr(normalization, 'invalid_count', 0)}`",
        f"- Skipped empty rows: `{getattr(normalization, 'skipped_empty', 0)}`",
        f"- Warehouse fact rows: `{getattr(warehouse, 'fact_rows', 0)}`",
        f"- DRE differences against Overview RCO: `{count_dre_differences(differences)}`",
        "",
        "## DRE Differences",
        "",
    ]
    if count_dre_differences(differences) == 0:
        lines.append("No DRE value differences were found.")
    else:
        lines.append("| company | division | period | DRE node | expected Overview value | calculated value | difference | status | probable business rule |")
        lines.append("|---|---|---|---|---:|---:|---:|---|---|")
        for row in differences[1:]:
            if row[7] != "PASS":
                lines.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | {row[8]} |")
    return "\n".join(lines) + "\n"


def write_pdf(path: Path, lines: list[str]) -> None:
    pages = [lines[i : i + 42] for i in range(0, len(lines), 42)] or [[]]
    objects: list[bytes] = []

    def add_object(payload: bytes) -> int:
        objects.append(payload)
        return len(objects)

    font_id = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    page_ids: list[int] = []
    content_ids: list[int] = []
    for page in pages:
        text = ["BT", "/F1 10 Tf", "50 790 Td", "14 TL"]
        for line in page:
            safe = line.encode("latin-1", "replace").decode("latin-1").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            text.append(f"({safe}) Tj")
            text.append("T*")
        text.append("ET")
        stream = "\n".join(text).encode("latin-1")
        content_ids.append(add_object(b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"))
        page_ids.append(0)
    pages_id = len(objects) + len(page_ids) + 1
    for index, content_id in enumerate(content_ids):
        page_payload = (
            f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 612 792] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>"
        ).encode("latin-1")
        page_ids[index] = add_object(page_payload)
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    pages_id = add_object(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("latin-1"))
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode("latin-1"))

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, payload in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("latin-1"))
        output.extend(payload)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    output.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("latin-1")
    )
    path.write_bytes(output)


def main() -> None:
    if SOURCE.resolve() != (ROOT / "data" / "master" / "Demonstrativos_Financeiros_2026_Rel_Razao_Active.xlsx").resolve():
        raise RuntimeError("Acceptance must use the official accounting workbook only")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for file_name in OUTPUT_FILES:
        path = OUTPUT_DIR / file_name
        if path.exists():
            path.unlink()

    source_hash = workbook_hash(SOURCE)
    report, warehouse_builder, provider, total_runtime = run_pipeline()
    if not report.success:
        raise RuntimeError(f"Pipeline failed: {report.errors}")

    store = warehouse_builder.store
    expected = overview_expected_values()
    calculated_rows, actual_by_code = calculated_dre_rows(report, provider)
    dre_differences = difference_rows(expected, actual_by_code, provider)
    reconciliation = reconciliation_rows(report, dre_differences)
    fato_dre = fato_dre_rows(report, provider)

    fact = fact_rows(store.facts.values())
    structural = structural_rows(report)
    dimensions = {
        "DIM_COMPANY": dimension_rows(store.companies.values()),
        "DIM_DIVISION": dimension_rows(store.divisions.values()),
        "DIM_COSTCENTER": dimension_rows(store.costcenters.values()),
        "DIM_ACCOUNT": dimension_rows(store.accounts.values()),
        "DIM_PERIOD": dimension_rows(store.periods.values()),
    }

    save_workbook(OUTPUT_DIR / "FACT_ACCOUNTING_ENTRY.xlsx", {"FACT_ACCOUNTING_ENTRY": fact})
    save_workbook(OUTPUT_DIR / "FATO_DRE.xlsx", {"FATO_DRE": fato_dre})
    save_workbook(OUTPUT_DIR / "ODS.xlsx", {**dimensions, "ODS_FACT": fact, "STRUCTURAL_ROWS": structural})
    save_workbook(OUTPUT_DIR / "Warehouse.xlsx", {"Summary": [
        ["metric", "value"],
        ["companies", len(store.companies)],
        ["divisions", len(store.divisions)],
        ["costcenters", len(store.costcenters)],
        ["accounts", len(store.accounts)],
        ["periods", len(store.periods)],
        ["facts", len(store.facts)],
        ["structural_rows", max(len(structural) - 1, 0)],
    ], **dimensions, "FACT_ACCOUNTING_ENTRY": fact, "STRUCTURAL_ROWS": structural})
    save_workbook(OUTPUT_DIR / "Calculated_DRE.xlsx", {"Calculated_DRE": calculated_rows})
    save_workbook(OUTPUT_DIR / "Reconciliation_Report.xlsx", {"Reconciliation": reconciliation})
    save_workbook(OUTPUT_DIR / "Validation_Report.xlsx", {"Validation_Report": dre_differences})

    metrics = {
        "source_workbook": str(SOURCE.relative_to(ROOT)),
        "source_sha256": source_hash,
        "pipeline_success": report.success,
        "total_runtime_seconds": total_runtime,
        "pipeline_duration_seconds": report.duration_seconds,
        "stages": [
            {
                "name": stage.name,
                "success": stage.success,
                "duration_seconds": stage.duration_seconds,
                "errors": stage.errors,
                "warnings": stage.warnings,
            }
            for stage in report.stage_results
        ],
        "counts": {
            "fact_rows": len(store.facts),
            "normalized_entries": getattr(report.metadata["workflow_payload"].get("normalization_report"), "normalized_count", 0),
            "invalid_rows": getattr(report.metadata["workflow_payload"].get("normalization_report"), "invalid_count", 0),
            "skipped_empty_rows": getattr(report.metadata["workflow_payload"].get("normalization_report"), "skipped_empty", 0),
            "rel_razao_rows_covered": len(store.facts) + max(len(structural) - 1, 0),
            "fato_dre_rows": max(len(fato_dre) - 1, 0),
            "dre_nodes": max(len(calculated_rows) - 1, 0),
            "dre_differences": count_dre_differences(dre_differences),
        },
    }
    (OUTPUT_DIR / "Execution_Metrics.json").write_text(json.dumps(serialise(metrics), ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")

    validation_md = validation_markdown(report, dre_differences, source_hash)
    (OUTPUT_DIR / "Validation_Report.md").write_text(validation_md, encoding="utf-8")

    pdf_lines = [
        "ARDO Financial Platform - Acceptance Executive Summary",
        "",
        f"Official workbook: {SOURCE.name}",
        f"Workbook SHA256: {source_hash}",
        f"Pipeline success: {report.success}",
        f"Runtime seconds: {total_runtime:.3f}",
        f"Fact rows: {len(store.facts)}",
        f"DRE nodes: {max(len(calculated_rows) - 1, 0)}",
        f"DRE differences vs Overview RCO: {count_dre_differences(dre_differences)}",
        "",
        "Stage results:",
    ]
    pdf_lines.extend([f"- {stage.name}: success={stage.success}, seconds={stage.duration_seconds:.3f}" for stage in report.stage_results])
    pdf_lines.extend(["", "DRE differences:"])
    if count_dre_differences(dre_differences) == 0:
        pdf_lines.append("- None")
    else:
        failed_rows = [row for row in dre_differences[1:] if row[7] != "PASS"]
        for row in failed_rows[:60]:
            pdf_lines.append(f"- {row[3]} | company={row[0]} | division={row[1]} | period={row[2]} | expected={row[4]} | calculated={row[5]} | diff={row[6]} | rule={row[8]}")
        if len(failed_rows) > 60:
            pdf_lines.append(f"- Additional differences omitted from PDF: {len(failed_rows) - 60}")
    write_pdf(OUTPUT_DIR / "Executive_Summary.pdf", pdf_lines)

    print(json.dumps({"output_dir": str(OUTPUT_DIR), "files": sorted(path.name for path in OUTPUT_DIR.iterdir())}, indent=2))


if __name__ == "__main__":
    main()
